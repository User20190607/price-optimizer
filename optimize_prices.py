#!/usr/bin/env python3
"""
单价调整优化器 v2.0  ——  GCD预检 + 四种算法 + MILP兜底
============================================================
模式说明
  uniform    (A) 等比均匀下浮 + 贪心残差分配（默认，投标场景）
  minimal    (B) ???? MILP ?????????
  single_rate(C) 单一折扣率 + 尾差平衡行（审计友好场景）
  milp           通用MILP兜底（复杂约束场景）

用法:
  python optimize_prices.py <excel> <target> [--mode uniform|minimal|single_rate|milp]
  python optimize_prices.py input.xlsx 9100 --mode minimal
  python optimize_prices.py input.xlsx 9100 --mode single_rate --balance-item "调整项"
"""

import sys, argparse, os, math
from decimal import Decimal, ROUND_HALF_UP
from functools import reduce

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 颜色常量 ─────────────────────────────────────────────────
CLR_TOTAL   = "D9E1F2"
CLR_TARGET  = "E2EFDA"
CLR_CHANGED = "FCE4D6"   # 模式B/C 橙色高亮
CLR_UNIFORM = "FFF2CC"   # 模式A 黄色高亮

# ══════════════════════════════════════════════════════════════
# 工具函数
# ══════════════════════════════════════════════════════════════

def safe_cents(value) -> int:
    """把元价格转成分（避免浮点误差），使用 Decimal 中间计算"""
    return int(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100)

def gcd2(a: int, b: int) -> int:
    a, b = abs(a), abs(b)
    while b:
        a, b = b, a % b
    return a

def list_gcd(nums):
    return reduce(gcd2, [int(round(n)) for n in nums])

def gcd_preflight(quantities_int, orig_total_cents, target_cents):
    """GCD 预检：返回 (g, snapped, gap, residual)，保证 gap 可被 g 整除"""
    g = reduce(gcd2, quantities_int)
    snapped = round(target_cents / g) * g
    return g, snapped, snapped - orig_total_cents, target_cents - snapped

def preflight(quantities, orig_prices, target_yuan: float) -> dict:
    """
    GCD 可行性预检。
    返回:
      granularity_yuan  可达步长（元）
      snapped_target    吸附后目标（分）
      gap_cents         吸附目标 - 原始总价（分）
      unavoidable_err   不可消除误差（元）
      exactly_reachable 是否精确可达
    """
    orig_total_cents = sum(q * safe_cents(p) for q, p in zip(quantities, orig_prices))
    target_cents = safe_cents(target_yuan)
    g, snapped, gap, residual = gcd_preflight(quantities, orig_total_cents, target_cents)
    return {
        "g_cents": g,
        "granularity_yuan": g / 100.0,
        "snapped_target": snapped,
        "gap_cents": gap,
        "unavoidable_err": abs(residual) / 100.0,
        "exactly_reachable": residual == 0,
    }


# ══════════════════════════════════════════════════════════════
# 模式A：等比均匀下浮 + 贪心残差分配
# ══════════════════════════════════════════════════════════════

def run_uniform(quantities, orig_prices, pre: dict) -> tuple:
    """
    所有单价等比缩放到目标，取整后按"大数量优先"贪心分配残差，
    保证合计精确等于 pre['snapped_target']。
    返回: (new_price_cents_list, changed_items, ratio)
    """
    n = len(quantities)
    orig_total_cents = sum(q * safe_cents(p) for q, p in zip(quantities, orig_prices))
    ratio = pre["snapped_target"] / orig_total_cents
    new_cents = [round(safe_cents(p) * ratio) for p in orig_prices]

    current_sum = sum(q * c for q, c in zip(quantities, new_cents))
    residual = pre["snapped_target"] - current_sum

    # 按数量从大到小分配残差（单价畸变最小化）
    sorted_idx = sorted(range(n), key=lambda i: -quantities[i])
    for i in sorted_idx:
        if residual == 0:
            break
        step = int(residual / quantities[i])
        if step != 0:
            new_cents[i] += step
            residual -= step * int(quantities[i])

    # 极小残差强制压到数量最大的行（已是 GCD 格点内，误差 ≤ g/qty 分）
    if residual != 0:
        new_cents[sorted_idx[0]] += int(residual / quantities[sorted_idx[0]])

    changed = []
    for i in range(n):
        diff = new_cents[i] - safe_cents(orig_prices[i])
        if diff != 0:
            changed.append({
                "index": i, "orig": orig_prices[i],
                "new": new_cents[i] / 100,
                "delta_cents": diff,
                "qty": quantities[i],
            })
    return new_cents, changed, ratio


# ══════════════════════════════════════════════════════════════
# 模式B：最少项数 MILP 精确解
# ══════════════════════════════════════════════════════════════
#
# 最小化被修改的行数本质是 L0 范数最小化（基数最小化），
# 整数/连续优化无法直接写 count(delta != 0)。标准做法是引入
# 0/1 指示变量 z_i，用 big-M 将其与 delta_i 联动：
#   -M_i * z_i <= delta_i <= M_i * z_i
# 目标 min sum(z_i) 即精确等于"改了几行"。
#
# big-M 取 |gap| 不是随手写的：单行独立吸收全部差额时，
# per-unit 调整为 gap/q_i，其绝对值 <= |gap|（因 q_i >= 1），
# 故 M_i = |gap| 既不会切掉最优解，又不会大到引发数值不稳。
#
# 松弛变量 s_pos/s_neg + big_penalty 实现"绝不返回 infeasible"：
# 正常（已过 GCD 预检）gap 可被 g 整除，松弛被压到 0；
# 若被绕过预检的脏数据触发，松弛吸收余数返回最小误差解。
# ══════════════════════════════════════════════════════════════

def minimal_items_milp(quantities, orig_cents, gap_cents,
                       max_items=None, max_change_cents=None,
                       big_penalty=10**6, msg=False):
    """
    最小化被修改的行数，精确凑平差额 gap_cents（分）。
    前置假设：已过 GCD 预检，gap_cents 可被 gcd(quantities) 整除。
    """
    import pulp

    n = len(quantities)
    prob = pulp.LpProblem("Minimal_Items", pulp.LpMinimize)

    delta = {i: pulp.LpVariable(f"d_{i}", cat="Integer") for i in range(n)}
    z     = {i: pulp.LpVariable(f"z_{i}", cat="Binary")  for i in range(n)}
    s_pos = pulp.LpVariable("s_pos", lowBound=0)
    s_neg = pulp.LpVariable("s_neg", lowBound=0)

    def M(i):
        m = abs(gap_cents)
        if max_change_cents is not None:
            m = min(m, max_change_cents)
        return max(m, 1)

    # 约束1：总差额平衡（带松弛兜底）
    prob += pulp.lpSum(quantities[i] * delta[i] for i in range(n)) + s_neg - s_pos == gap_cents

    # 约束2：big-M 联动 -- delta_i != 0 => z_i = 1
    for i in range(n):
        prob += delta[i] <=  M(i) * z[i]
        prob += delta[i] >= -M(i) * z[i]
    # 约束3：新单价 >= 1 分钱
        prob += delta[i] >= -orig_cents[i] + 1
    # 约束4（可选）：最多修改项数
    if max_items is not None:
        prob += pulp.lpSum(z[i] for i in range(n)) <= max_items

    # 目标：最小化被改行数，松弛以大惩罚兜底
    prob += big_penalty * (s_pos + s_neg) + pulp.lpSum(z[i] for i in range(n))
    prob.solve(pulp.PULP_CBC_CMD(msg=msg))

    new_cents = [orig_cents[i] + int(round(delta[i].value() or 0)) for i in range(n)]
    changed = [{"index": i, "orig": orig_cents[i]/100, "new": new_cents[i]/100,
                "delta_cents": new_cents[i] - orig_cents[i], "qty": quantities[i]}
               for i in range(n) if new_cents[i] != orig_cents[i]]
    leftover = round((s_pos.value() or 0) - (s_neg.value() or 0))
    return {"status": pulp.LpStatus[prob.status],
            "new_prices_cents": new_cents,
            "changed_items": changed,
            "num_changed": len(changed),
            "leftover_cents": leftover}

def run_minimal(quantities, orig_prices, pre: dict, max_change_cents=None) -> tuple:
    """
    模式B入口：调用 minimal_items_milp，返回兼容的
    (new_cents_list, changed_items, num_changed)。
    """
    orig_cents = [safe_cents(p) for p in orig_prices]
    gap = pre["gap_cents"]

    if gap == 0:
        return orig_cents, [], 0

    result = minimal_items_milp(
        quantities, orig_cents, gap,
        max_change_cents=max_change_cents,
        msg=False,
    )

    if result["leftover_cents"] != 0:
        print(f"Warning: MILP slack {result['leftover_cents']} fen (post-GCD, expect 0)")

    return result["new_prices_cents"], result["changed_items"], result["num_changed"]

# ══════════════════════════════════════════════════════════════
# 模式C：单一折扣率 + 平衡行（审计友好）
# ══════════════════════════════════════════════════════════════

def run_single_rate(quantities, orig_prices, pre: dict,
                    balance_name="调整尾差项") -> tuple:
    """
    用一个统一折扣率（保留4位小数）对所有行计算新单价，
    取整后把残差放到一行专用"尾差调整"项（加在表末）。
    返回: (new_price_cents_list, balance_item, rate, total_check)
    balance_item = {"name": str, "qty": 1, "unit_price_cents": int}
    """
    orig_total_cents = sum(q * safe_cents(p) for q, p in zip(quantities, orig_prices))
    rate = round(pre["snapped_target"] / orig_total_cents, 6)  # 统一折扣率

    new_cents = [round(safe_cents(p) * rate) for p in orig_prices]
    sub_total = sum(q * c for q, c in zip(quantities, new_cents))
    balance_cents = pre["snapped_target"] - sub_total  # 尾差（分）

    total_check = sub_total + balance_cents
    return new_cents, {
        "name": balance_name,
        "qty": 1,
        "unit_price_cents": balance_cents,
    }, rate, total_check


# ══════════════════════════════════════════════════════════════
# MILP 通用兜底（需要 pulp）
# ══════════════════════════════════════════════════════════════

def run_milp_uniform(quantities, orig_prices, pre: dict, max_delta_cents=5) -> tuple:
    """原始 MILP 均匀模式，作为复杂约束兜底"""
    try:
        import pulp
    except ImportError:
        raise ImportError("MILP 模式需要安装 pulp：pip install pulp")

    n = len(quantities)
    orig_total_cents = sum(q * safe_cents(p) for q, p in zip(quantities, orig_prices))
    ratio = pre["snapped_target"] / orig_total_cents
    theory_cents = [round(safe_cents(p) * ratio) for p in orig_prices]

    prob = pulp.LpProblem("MILP_Uniform", pulp.LpMinimize)
    x = {i: pulp.LpVariable(f"P_{i}",
                             lowBound=theory_cents[i] - max_delta_cents,
                             upBound=theory_cents[i] + max_delta_cents,
                             cat='Integer') for i in range(n)}
    e_pos = pulp.LpVariable("Ep", lowBound=0, cat='Continuous')
    e_neg = pulp.LpVariable("En", lowBound=0, cat='Continuous')
    d = {i: pulp.LpVariable(f"D_{i}", lowBound=0, cat='Continuous') for i in range(n)}

    prob += pulp.lpSum(quantities[i] * x[i] for i in range(n)) + e_neg - e_pos \
           == pre["snapped_target"]
    for i in range(n):
        prob += d[i] >= x[i] - theory_cents[i]
        prob += d[i] >= theory_cents[i] - x[i]

    M = 100000
    prob += M * (e_pos + e_neg) + pulp.lpSum(d[i] for i in range(n))
    prob.solve(pulp.PULP_CBC_CMD(msg=False))

    if pulp.LpStatus[prob.status] != 'Optimal':
        if max_delta_cents < 200:
            return run_milp_uniform(quantities, orig_prices, pre, max_delta_cents * 2)
        raise ValueError("MILP 无法在合理范围内找到可行解")

    new_cents = [int(round(x[i].varValue)) for i in range(n)]
    changed = []
    for i in range(n):
        diff = new_cents[i] - theory_cents[i]
        if diff != 0:
            changed.append({"index": i, "orig": orig_prices[i],
                             "new": new_cents[i] / 100, "delta_cents": diff,
                             "qty": quantities[i]})
    return new_cents, changed, ratio


# ══════════════════════════════════════════════════════════════
# 列名自动识别
# ══════════════════════════════════════════════════════════════

def detect_columns(df) -> dict:
    col_map = {}
    for col in df.columns:
        nl = str(col).strip().lower()
        if any(k in nl for k in ['规格', '型号', '名称', 'spec', 'item', '材料', '描述', '品名']):
            col_map.setdefault('spec', col)
        elif any(k in nl for k in ['数量', 'qty', 'quantity', '用量']):
            col_map.setdefault('qty', col)
        elif any(k in nl for k in ['单价', 'unit price', 'price', '综合单价']):
            col_map.setdefault('price', col)
        elif any(k in nl for k in ['合价', '金额', '小计', '合计', 'total', 'amount']):
            col_map.setdefault('total', col)
    return col_map


# ══════════════════════════════════════════════════════════════
# Excel 读取 + 写回（四种模式共用）
# ══════════════════════════════════════════════════════════════

def process_excel(input_path, target_total, output_path=None, mode='uniform',
                  max_delta_cents=5, max_delta_yuan=None, max_items=None,
                  balance_item_name="调整尾差项",
                  spec_col=None, qty_col=None, price_col=None, total_col=None,
                  header_row=0, sheet_name=0):

    df = pd.read_excel(input_path, sheet_name=sheet_name, header=header_row)
    df.columns = df.columns.map(str)

    col_map = detect_columns(df)
    if spec_col:  col_map['spec']  = spec_col
    if qty_col:   col_map['qty']   = qty_col
    if price_col: col_map['price'] = price_col
    if total_col: col_map['total'] = total_col

    missing = [k for k in ['qty', 'price'] if k not in col_map]
    if missing:
        raise ValueError(
            f"无法自动识别列 {missing}，请用 --qty-col / --price-col 手动指定。\n"
            f"已有列: {list(df.columns)}"
        )

    qty_c, price_c = col_map['qty'], col_map['price']
    mask = (pd.to_numeric(df[qty_c], errors='coerce').notna() &
            pd.to_numeric(df[price_c], errors='coerce').notna())
    data_df = df[mask].copy()
    data_df[qty_c]   = pd.to_numeric(data_df[qty_c])
    data_df[price_c] = pd.to_numeric(data_df[price_c])
    data_df = data_df[data_df[qty_c] > 0]

    if len(data_df) == 0:
        raise ValueError("未找到有效数据行（数量>0 且单价为数字）")

    quantities  = data_df[qty_c].tolist()
    orig_prices = data_df[price_c].tolist()
    specs       = (data_df[col_map['spec']].tolist()
                   if 'spec' in col_map else [f"Item_{i}" for i in range(len(quantities))])
    orig_total  = sum(q * p for q, p in zip(quantities, orig_prices))

    # ── GCD 预检 ──────────────────────────────────────────────
    pre = preflight(quantities, orig_prices, target_total)
    print(f"\n{'─'*60}")
    print(f"  GCD 预检结果")
    print(f"  可达步长: {pre['granularity_yuan']:.4f} 元")
    if not pre['exactly_reachable']:
        print(f"  ⚠️  目标不在格点，自动吸附到 {pre['snapped_target']/100:.4f} 元")
        print(f"     不可消除误差: {pre['unavoidable_err']:.4f} 元")
    else:
        print(f"  ✅ 目标精确可达")
    print(f"{'─'*60}")

    # ── 调用对应模式 ──────────────────────────────────────────
    balance_item = None
    ratio = None
    num_changed = 0
    mode_label = ""

    if mode == 'uniform':
        new_cents, changed, ratio = run_uniform(quantities, orig_prices, pre)
        mode_label = "模式A：等比均匀下浮 + 贪心残差"
        num_changed = len(changed)

    elif mode == 'minimal':
        max_change_cents = int(max_delta_yuan * 100) if max_delta_yuan is not None else None
        new_cents, changed, num_changed = run_minimal(
            quantities, orig_prices, pre, max_change_cents)
        mode_label = "模式B：最少项数 MILP 精确解"

    elif mode == 'single_rate':
        new_cents, balance_item, ratio, _ = run_single_rate(
            quantities, orig_prices, pre, balance_item_name)
        changed = []  # 单一费率每行都变，不单独标记
        num_changed = len(quantities)
        mode_label = "模式C：单一折扣率 + 平衡行"

    elif mode == 'milp':
        new_cents, changed, ratio = run_milp_uniform(
            quantities, orig_prices, pre, max_delta_cents)
        mode_label = "MILP 通用兜底"
        num_changed = len(changed)

    else:
        raise ValueError(f"未知模式: {mode}")

    new_total_cents = (sum(int(round(quantities[i])) * new_cents[i] for i in range(len(quantities)))
                       + (balance_item['unit_price_cents'] if balance_item else 0))
    new_total = new_total_cents / 100.0

    # ── 输出路径 ──────────────────────────────────────────────
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        suffix_map = {'uniform': '_均匀下浮', 'minimal': '_最少调整',
                      'single_rate': '_单一费率', 'milp': '_MILP'}
        output_path = base + suffix_map.get(mode, '_优化') + ext

    # ── 写回 Excel ────────────────────────────────────────────
    wb = load_workbook(input_path)
    ws = (wb.worksheets[sheet_name]
          if isinstance(sheet_name, int) else wb[sheet_name])

    excel_header_row = header_row + 1
    price_col_idx = df.columns.get_loc(price_c) + 1  # 1-based
    insert_at = price_col_idx + 1
    ws.insert_cols(insert_at, 2)
    new_price_letter = get_column_letter(insert_at)
    rate_letter      = get_column_letter(insert_at + 1)

    hfont = Font(bold=True, color="FFFFFF")
    for letter, text, color in [
        (new_price_letter, "优化单价", "2E75B6"),
        (rate_letter, "变动(元)", "375623" if mode in ('uniform', 'single_rate') else "7B2C2C"),
    ]:
        cell = ws[f"{letter}{excel_header_row}"]
        cell.value = text
        cell.font = hfont
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal='center')

    data_indices = data_df.index.tolist()
    changed_idx_set = {c['index'] for c in changed}
    fill_changed = PatternFill("solid", fgColor=CLR_CHANGED if mode in ('minimal', 'single_rate') else CLR_UNIFORM)

    for k, orig_idx in enumerate(data_indices):
        erow = orig_idx + header_row + 2
        new_p = new_cents[k] / 100.0
        orig_p = orig_prices[k]
        changed_flag = (mode == 'single_rate') or (k in changed_idx_set)

        cell_np = ws[f"{new_price_letter}{erow}"]
        cell_np.value = round(new_p, 2)
        cell_np.number_format = '#,##0.00'

        cell_d = ws[f"{rate_letter}{erow}"]
        cell_d.value = round(new_p - orig_p, 2)
        cell_d.number_format = '+#,##0.00;-#,##0.00;"-"'

        if changed_flag:
            cell_np.fill = fill_changed
            cell_d.fill = fill_changed

    ws.column_dimensions[new_price_letter].width = 12
    ws.column_dimensions[rate_letter].width = 12

    # 模式C：追加平衡行
    if balance_item:
        last_row = max(i + header_row + 2 for i in data_indices) + 1
        ws.cell(row=last_row, column=1, value=balance_item['name']).font = Font(bold=True, color="FF0000")
        qty_col_idx = df.columns.get_loc(qty_c) + 1
        ws.cell(row=last_row, column=qty_col_idx, value=1)
        ws.cell(row=last_row, column=price_col_idx, value=0)  # 原单价占位
        ws.cell(row=last_row, column=insert_at, value=round(balance_item['unit_price_cents'] / 100, 2))
        ws.cell(row=last_row, column=insert_at).fill = PatternFill("solid", fgColor="FF6600")
        ws.cell(row=last_row, column=insert_at).font = Font(bold=True, color="FFFFFF")

    # ── 汇总行 ────────────────────────────────────────────────
    last_data_row = max(i + header_row + 2 for i in data_indices)
    sr = last_data_row + (3 if balance_item else 2)
    sfill = PatternFill("solid", fgColor=CLR_TOTAL)
    tfill = PatternFill("solid", fgColor=CLR_TARGET)

    def sumrow(row, label, value, fill, fmt='#,##0.00'):
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(bold=True)
        c1.fill = fill
        cv = ws.cell(row=row, column=insert_at, value=value)
        cv.number_format = fmt
        cv.font = Font(bold=True)
        cv.fill = fill

    sumrow(sr,   "📊 原始总价（元）",   orig_total,   sfill)
    sumrow(sr+1, "🎯 目标总价（元）",   target_total, tfill)
    sumrow(sr+2, "✅ 优化后总价（元）", new_total,    tfill)
    sumrow(sr+3, "⚖️ 误差（元）",       abs(new_total - target_total), sfill, '#,##0.0000')
    sumrow(sr+4, "📉 整体下浮率",       (new_total - orig_total) / orig_total, sfill, '0.000%')
    sumrow(sr+5, "✏️ 被调整项数",       num_changed,  sfill, '0')
    if ratio:
        sumrow(sr+6, "📐 折扣系数",     ratio, sfill, '0.000000')

    wb.save(output_path)

    # ── 控制台报告 ────────────────────────────────────────────
    sep = '=' * 60
    print(f"\n{sep}")
    print(f"  {mode_label}")
    print(f"{sep}")
    print(f"  原始总价:    {orig_total:>14,.2f} 元")
    print(f"  目标总价:    {target_total:>14,.2f} 元")
    print(f"  优化总价:    {new_total:>14,.2f} 元")
    print(f"  绝对误差:    {abs(new_total-target_total):>14.4f} 元")
    print(f"  整体下浮:    {(new_total-orig_total)/orig_total*100:>14.3f} %")
    print(f"  被调整项数:  {num_changed:>14} 项")
    if ratio:
        print(f"  折扣系数:    {ratio:>14.6f}")
    print(f"{sep}")

    if changed:
        color_label = "橙色" if mode in ('minimal',) else "黄色"
        print(f"\n{color_label}高亮明细（{len(changed)} 项）:")
        for c in changed:
            print(f"  • {specs[c['index']][:30]}: "
                  f"{c['orig']:.2f} → {c['new']:.2f}  "
                  f"({'↑' if c['delta_cents']>0 else '↓'}{abs(c['delta_cents'])/100:.2f}元/件"
                  f" × {c['qty']} 件 = {abs(c['delta_cents'])*c['qty']/100:.2f}元)")
    if balance_item:
        print(f"\n平衡行: [{balance_item['name']}]  {balance_item['unit_price_cents']/100:.2f} 元 (×1)")

    print(f"\n📁 已保存: {output_path}\n")
    return output_path, new_total, abs(new_total - target_total)


# ══════════════════════════════════════════════════════════════
# CLI 入口
# ══════════════════════════════════════════════════════════════

def main():
    p = argparse.ArgumentParser(description='单价调整优化器 v2.0')
    p.add_argument('input',  help='输入Excel路径')
    p.add_argument('target', type=float, help='目标总价（元）')
    p.add_argument('-o', '--output', help='输出路径（默认自动命名）')
    p.add_argument('--mode',
                   choices=['uniform', 'minimal', 'single_rate', 'milp'],
                   default='uniform',
                   help='uniform=均匀下浮(默认) minimal=最少项数 single_rate=单一费率 milp=MILP兜底')
    p.add_argument('--max-delta',      type=int,   default=5,    help='[milp] 单价最大偏差(分)，默认5')
    p.add_argument('--max-change-per-item', type=float, default=None,
                   help='[minimal] 单项最大调整金额（元），默认不限')
    p.add_argument('--max-items',      type=int,   default=None, help='[minimal] 最多改几项')
    p.add_argument('--balance-item',   default='调整尾差项',     help='[single_rate] 平衡行名称')
    p.add_argument('--sheet',          default=0)
    p.add_argument('--header-row',     type=int,   default=0)
    p.add_argument('--spec-col')
    p.add_argument('--qty-col')
    p.add_argument('--price-col')
    p.add_argument('--total-col')
    args = p.parse_args()

    sheet = int(args.sheet) if str(args.sheet).isdigit() else args.sheet

    process_excel(
        input_path=args.input,
        target_total=args.target,
        output_path=args.output,
        mode=args.mode,
        max_delta_cents=args.max_delta,
        max_delta_yuan=args.max_change_per_item,
        max_items=args.max_items,
        balance_item_name=args.balance_item,
        spec_col=args.spec_col,
        qty_col=args.qty_col,
        price_col=args.price_col,
        total_col=args.total_col,
        header_row=args.header_row,
        sheet_name=sheet,
    )

if __name__ == '__main__':
    main()
